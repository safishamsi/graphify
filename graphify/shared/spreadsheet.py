"""Excel/CSV extraction → shared Table structure."""
from __future__ import annotations

import csv
import re
from pathlib import Path

from graphify.shared.tables import Table


def extract_excel_tables(path: Path) -> list[Table]:
    """Each sheet → one or more Table objects."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    tables = []

    for sheet in wb.worksheets:
        # Use merged-cell-aware extraction
        rows = _expand_merged(sheet)

        # Skip empty sheets
        if not rows or all(all(c == "" for c in r) for r in rows):
            continue

        # Detect sub-tables within a single sheet (split on blank rows)
        for region_idx, region in enumerate(_split_regions(rows)):
            if len(region) < 2:
                continue
            headers, body = _detect_header(region)
            table_id = f"{path.stem}__{sheet.title}"
            if region_idx > 0:
                table_id += f"__region_{region_idx}"
            tables.append(Table(
                id=table_id,
                caption=sheet.title,
                headers=headers,
                rows=body,
                source_file=str(path),
            ))

    wb.close()
    return tables


def extract_csv_table(path: Path) -> list[Table]:
    """Single CSV file → one Table."""
    rows = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append([c.strip() for c in row])

    if len(rows) < 2:
        return []

    headers, body = _detect_header(rows)
    return [Table(
        id=path.stem,
        caption=path.name,
        headers=headers,
        rows=body,
        source_file=str(path),
    )]


def _expand_merged(sheet) -> list[list[str]]:
    """Read sheet respecting merged cell ranges."""
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    grid = [[""] * max_col for _ in range(max_row)]

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[cell.row - 1][cell.column - 1] = _cell_str(cell.value)

    # Fill merged ranges with top-left value
    for merged in sheet.merged_cells.ranges:
        val = grid[merged.min_row - 1][merged.min_col - 1]
        for r in range(merged.min_row - 1, merged.max_row):
            for c in range(merged.min_col - 1, merged.max_col):
                grid[r][c] = val

    return grid


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _detect_header(rows: list[list[str]]) -> tuple[list[list[str]], list[list[str]]]:
    """Heuristic: first row is header if it's all text and rest has numbers."""
    if not rows:
        return [], []
    first = rows[0]
    rest = rows[1:]

    numeric_in_first = sum(1 for c in first if _is_numeric(c))
    numeric_in_rest = sum(1 for r in rest for c in r if _is_numeric(c))

    if numeric_in_first == 0 and numeric_in_rest > 0:
        return [first], rest
    return [], rows


def _is_numeric(s: str) -> bool:
    s = s.replace(",", "").replace("$", "").replace("%", "").strip()
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _split_regions(rows: list[list[str]]) -> list[list[list[str]]]:
    """Split a sheet into sub-tables separated by blank rows."""
    regions: list[list[list[str]]] = []
    current: list[list[str]] = []
    for row in rows:
        if all(c == "" for c in row):
            if current:
                regions.append(current)
                current = []
        else:
            current.append(row)
    if current:
        regions.append(current)
    return regions if regions else [rows]
