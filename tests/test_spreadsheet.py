"""Tests for Excel/CSV table extraction."""
import csv
from pathlib import Path

import pytest

from graphify.shared.spreadsheet import extract_csv_table, _detect_header, _split_regions


class TestCSV:
    def test_basic_csv(self, tmp_path):
        csv_file = tmp_path / "data.csv"
        csv_file.write_text("Name,Amount,Year\nAcme,100,2024\nBeta,200,2023\n")
        tables = extract_csv_table(csv_file)
        assert len(tables) == 1
        t = tables[0]
        assert t.headers == [["Name", "Amount", "Year"]]
        assert t.rows == [["Acme", "100", "2024"], ["Beta", "200", "2023"]]

    def test_csv_too_short(self, tmp_path):
        csv_file = tmp_path / "short.csv"
        csv_file.write_text("only,one,row\n")
        tables = extract_csv_table(csv_file)
        assert len(tables) == 0


class TestDetectHeader:
    def test_text_header_numeric_body(self):
        rows = [["Name", "Revenue", "Growth"], ["Acme", "100", "5%"], ["Beta", "200", "10%"]]
        headers, body = _detect_header(rows)
        assert headers == [["Name", "Revenue", "Growth"]]
        assert len(body) == 2

    def test_all_text(self):
        rows = [["A", "B"], ["C", "D"]]
        headers, body = _detect_header(rows)
        assert headers == []
        assert body == rows

    def test_empty(self):
        headers, body = _detect_header([])
        assert headers == []
        assert body == []


class TestSplitRegions:
    def test_single_block(self):
        rows = [["a", "1"], ["b", "2"]]
        regions = _split_regions(rows)
        assert len(regions) == 1

    def test_split_on_blank(self):
        rows = [["a", "1"], ["", ""], ["b", "2"]]
        regions = _split_regions(rows)
        assert len(regions) == 2
        assert regions[0] == [["a", "1"]]
        assert regions[1] == [["b", "2"]]

    def test_multiple_blanks(self):
        rows = [["a", "1"], ["", ""], ["", ""], ["b", "2"]]
        regions = _split_regions(rows)
        assert len(regions) == 2


class TestExcel:
    def test_excel_extraction(self, tmp_path):
        """Test Excel extraction if openpyxl is available."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from graphify.shared.spreadsheet import extract_excel_tables

        xlsx_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue"
        ws.append(["Company", "Revenue", "Year"])
        ws.append(["Acme", 1000000, 2024])
        ws.append(["Beta", 2000000, 2023])
        wb.save(xlsx_path)

        tables = extract_excel_tables(xlsx_path)
        assert len(tables) == 1
        t = tables[0]
        assert t.caption == "Revenue"
        assert t.headers == [["Company", "Revenue", "Year"]]
        assert len(t.rows) == 2
        assert t.rows[0][0] == "Acme"

    def test_excel_empty_sheet_skipped(self, tmp_path):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from graphify.shared.spreadsheet import extract_excel_tables

        xlsx_path = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)

        tables = extract_excel_tables(xlsx_path)
        assert len(tables) == 0

    def test_excel_multiple_regions(self, tmp_path):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from graphify.shared.spreadsheet import extract_excel_tables

        xlsx_path = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append(["1", "2"])
        ws.append([None, None])  # blank row separator
        ws.append(["X", "Y"])
        ws.append(["3", "4"])
        wb.save(xlsx_path)

        tables = extract_excel_tables(xlsx_path)
        assert len(tables) == 2
